import openpyxl


class Analysis(object):

    def __init__(self, class_name: str):
        self.__class_name = class_name
        try:
            self.__workbook_path = "Excel Files/" + class_name + ".xlsx"
        except FileNotFoundError:
            pass
            # TODO: Add code to handle the error

    def month_analysis(self, month: str) -> dict:
        """
        Returns analysis of student attendance of a given month
        :param month: month to analyse
        :return student_record: a dictionary containing a list of student records & average percentage of daily attendance for the month
        """
        wb = openpyxl.load_workbook(self.__workbook_path)
        student_record = list()

        if month not in wb.sheetnames:
            pass
            # TODO: Add code to handle invalid month
        else:
            wb._active_sheet_index = wb.sheetnames.index(month)
            month_sheet = wb.active
            row_count = 2
            student_name = month_sheet.cell(row_count, 1).value
            while student_name is not None:
                student = dict()
                student['name'] = student_name
                for date in range(31):
                    if month_sheet.cell(1, date+1) is not None:
                        temp = month_sheet.cell(row_count, date + 1)
                        if temp is None:
                            pass  # Student was not registered in the class yet
                        else:
                            if temp == 'P':
                                try:
                                    student['present'] += 1
                                except KeyError:
                                    student['present'] = 1
                            elif temp == 'H':
                                try:
                                    student['partial'] += 1
                                except KeyError:
                                    student['partial'] = 1
                            elif temp == 'A':
                                try:
                                    student['absent'] += 1
                                except KeyError:
                                    student['absent'] = 1
                            try:
                                student['total'] += 1
                            except KeyError:
                                student['total'] = 1
                student_record.append(student.copy())
                row_count += 1
                student_name = month_sheet.cell(row_count, 1).value
        total_present = 0
        total_days = 0
        for student in student_record:
            total_present += student['present']
            total_days = max(total_days, student['total'])

        wb.close()
        return {'record': student_record, 'average': (total_present/total_days)/len(student_record)*100}

    def duration_analysis(self, start_month: str, end_month: str) -> dict:
        """
        Returns analysis of student attendance of for the given duration
        :param start_month: string
        :param end_month: string
        :return: a dictionary containing a list of student records & average percentage of daily attendance for the entire duration
        """
        wb = openpyxl.load_workbook(self.__workbook_path)
        student_record = dict()
        if start_month in wb.sheetnames and end_month in wb.sheetnames and wb.sheetnames.index(start_month) <= wb.sheetnames.index(end_month):
            months = list()
            month_range = wb.sheetnames[wb.sheetnames.index(start_month), wb.sheetnames.index(end_month)+1]
            wb.close()

            for month in month_range:
                months.append(self.month_analysis(month))

            for month in months:
                for student in month['record']:
                    if student['name'] not in student_record.keys():
                        student_record['name'] = student.copy()
                    else:
                        student_record['name']['present'] += student['present']
                        student_record['name']['partial'] += student['partial']
                        student_record['name']['absent'] += student['absent']
                        student_record['name']['total'] += student['total']

            student_record = student_record.values()

            total_present = 0
            total_days = 0
            for student in student_record:
                total_present += student['present']
                total_days = max(total_days, student['total'])
            return {'record': student_record, 'average': (total_present / total_days) / len(student_record) * 100}
        else:
            pass
            # TODO: Error-> Invalid month range

if __name__ == '__main__':
    pass
