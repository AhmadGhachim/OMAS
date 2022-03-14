# external libs/modules
import os
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill
# local modules
import MathFunctions
from FileParser import FileParser
import ExcelFunctions

original_months = ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']

class DataProcessor(object):
    def __init__(self, workbook_path: str, meeting_file_path: str, meeting_service: str, class_duration: int, cutoff: int = 0):
        # Class details
        self.__class_name = meeting_file_path.split('/')[-1].split('.')[0]
        self.__class_duration = class_duration
        self.__class_list = list()

        # Meeting data
        self.__file_parser = FileParser(meeting_file_path)
        self.__meeting_service = meeting_service
        self.__cutoff = cutoff

        # Processed data
        self.__workbook_path = workbook_path
        self.__report = dict()
        self.__duration_db = dict()
        self.__error_db = list()

        self.call_processor(self.__meeting_service)
        self.process()

    def __populate_class_list(self):
        """
        Helper function for 'process'
        :return: None
        """
        wb = openpyxl.load_workbook(self.__workbook_path)  # Open class workbook
        sheet = wb.active

        # Create a list of student names from the workbook
        row_count = 2
        cell_data = sheet.cell(row_count, 1).value
        while cell_data is not None:
            self.__class_list.append(cell_data)
            row_count += 1
            cell_data = sheet.cell(row_count, 1).value

        wb.close()

    def modify_cutoff(self, new_cutoff: int):
        self.__cutoff = new_cutoff

    def call_processor(self, meeting_service: str):
        """
        Call parse methods from FileParser based on the argument
        :param meeting_service: string representing the name of the meeting service
        :return: None
        """
        if meeting_service == 'teams':
            self.__file_parser.parse_for_teams()
        elif meeting_service == 'zoom':
            self.__file_parser.parse_for_zoom()
        elif meeting_service == 'webex':
            self.__file_parser.parse_for_webex()
        else:
            raise ValueError("Specified meeting service is currently not supported")

    def process(self):
        """
        Uses data from the FileParser to generate serializable reports
        :return: None
        """
        self.__populate_class_list()

        # Check if all names are correctly entered
        exceptions = list()
        exceptions_replacement = list()

        i = 0
        for y in [z[0] for z in self.__file_parser.data]:
            if y in self.__class_list:
                pass
            else:
                exceptions.append(y)
                similarity, index = 0, 0
                count = 0
                for z in self.__class_list:
                    match = SequenceMatcher(a=y, b=z).ratio()
                    if match > similarity:
                        similarity = match
                        index = count
                    count += 1
                self.__file_parser.data[i][0] = self.__class_list[index]
                exceptions_replacement.append(self.__class_list[index])
            i += 1

        self.__error_db = [exceptions, exceptions_replacement]

        self.__report['present'] = list()
        self.__report['partial'] = list()
        self.__report['absent'] = list()

        for y in self.__file_parser.data:
            if y[0] not in self.__duration_db.keys():

                self.__duration_db[y[0]] = [y[2]['time']]
            else:
                self.__duration_db[y[0]].append(y[2]['time'])
        for y in self.__duration_db:
            if len(self.__duration_db[y]) % 2 == 1:
                self.__duration_db[y].append(MathFunctions.timestamp_from_duration(self.__file_parser.start_time, self.__class_duration))
        for y in self.__duration_db:
            self.__duration_db[y] = MathFunctions.total_duration(self.__duration_db[y])


        self.__report['absent'] = self.__class_list.copy()

        for y in self.__duration_db:
            self.__report['absent'].remove(y)
            if self.__duration_db[y] > self.__cutoff:
                self.__report['present'].append(y)
            elif self.__cutoff > 0:
                self.__report['partial'].append(y)

    def output_to_workbook(self):
        """
        Method which updates the Excel workbook database using the parsed data
        :return: None
        """
        # Styles
        present_text = Font(color='FF006D08')
        partial_text = Font(color='FFA45300')
        absent_text = Font(color='FFBE0020')

        present_style = PatternFill(start_color='FFBAF4CE', end_color='FFBAF4CE', fill_type='solid')
        partial_style = PatternFill(start_color='FFFFEE97', end_color='FFFFEE97', fill_type='solid')
        absent_style = PatternFill(start_color='FFFFC1CC', end_color='FFFFC1CC', fill_type='solid')

        wb = openpyxl.load_workbook(self.__workbook_path)
        wb._active_sheet_index = wb.sheetnames.index(original_months[int(self.__file_parser.date.split("/")[1])-1])  # Select month sheet
        sheet = wb.active

        # Write date on top of the column where data is to be stored
        sheet[ExcelFunctions.row_col_2_cell(1, int(self.__file_parser.date.split("/")[0]) + 1)] = \
            original_months[int(self.__file_parser.date.split("/")[1]) - 1][:3] + " " + \
            self.__file_parser.date.split("/")[0]

        # First mark everyone as absent
        for x in range(len(self.__class_list)):
            sheet[ExcelFunctions.row_col_2_cell(x + 2, int(self.__file_parser.date.split("/")[0]) + 1)] = 'A'
            sheet[ExcelFunctions.row_col_2_cell(x + 2, int(self.__file_parser.date.split("/")[0]) + 1)].font = absent_text
            sheet[ExcelFunctions.row_col_2_cell(x + 2, int(self.__file_parser.date.split("/")[0]) + 1)].fill = absent_style

        # Then mark students as present or partial
        for x in self.__report['present']:
            sheet[ExcelFunctions.row_col_2_cell(self.__class_list.index(x) + 2, int(self.__file_parser.date.split("/")[0]) + 1)] = "P"
            sheet[ExcelFunctions.row_col_2_cell(self.__class_list.index(x) + 2, int(self.__file_parser.date.split("/")[0]) + 1)].font = present_text
            sheet[ExcelFunctions.row_col_2_cell(self.__class_list.index(x) + 2, int(self.__file_parser.date.split("/")[0]) + 1)].fill = present_style

        for x in self.__report['partial']:
            sheet[ExcelFunctions.row_col_2_cell(self.__class_list.index(x) + 2, int(self.__file_parser.date.split("/")[0]) + 1)] = "H"
            sheet[ExcelFunctions.row_col_2_cell(self.__class_list.index(x) + 2, int(self.__file_parser.date.split("/")[0]) + 1)].font = partial_text
            sheet[ExcelFunctions.row_col_2_cell(self.__class_list.index(x) + 2, int(self.__file_parser.date.split("/")[0]) + 1)].fill = partial_style

        wb.save(self.__workbook_path)
        wb.close()

    def output_to_text_file(self):
        """
        Method which writes processed data to an external .txt file under 'Reports' directory
        :return: None
        """
        if "Reports" not in os.listdir():
            os.mkdir("Reports")
        report_file = open("Reports/" + self.__class_name + " report.txt", "w")
        report_file.write("Report for " + self.__class_name + " (" + self.__file_parser.date + ")\n")
        report_file.write("\nPresent: " + str(len(self.__report["present"])) + "\n")
        counter = 1
        for y in self.__report['present']:
            report_file.write("\t" + str(counter) + ". " + y + " (time: " + str(abs(int(self.__duration_db[y]))) + " mins)\n")
            counter += 1

        if len(self.__report['partial']):
            report_file.write("\nPartial (< " + str(self.__cutoff) + " mins): " + str(len(self.__report["partial"])) + "\n")
            counter = 1
            for y in self.__report['partial']:
                report_file.write("\t" + str(counter) + ". " + y + " (time: " + str(abs(int(self.__duration_db[y]))) + " mins)\n")
                counter += 1

        report_file.write("\nAbsent: " + str(len(self.__report["absent"])) + "\n")
        counter = 1
        for y in self.__report['absent']:
            report_file.write("\t" + str(counter) + ". " + y + "\n")
            counter += 1

        report_file.write("\nPotential Errors (spelling mistakes, etc.): " + str(len(self.__error_db[0])) + "\n")
        counter = 1
        for y in range(len(self.__error_db[0])):
            report_file.write("\t" + str(counter) + ". " + self.__error_db[0][y] + " (replaced with " + self.__error_db[1][y] + ")\n")
            counter += 1

        report_file.close()

    def output_to_console(self):
        """
        Method which writes processed data to console
        :return: None
        """
        print("Report for", self.__class_name, "(" + self.__file_parser.date + ")\n")
        print("Present: " + str(len(self.__report["present"])))
        counter = 1
        for y in self.__report['present']:
            print("\t" + str(counter) + ". " + y +" (time:", str(abs(int(self.__duration_db[y]))), "mins)")
            counter += 1

        print("\nPartial (<", self.__cutoff, "mins): " + str(len(self.__report["partial"])))
        counter = 1
        for y in self.__report['partial']:
            print("\t" + str(counter) + ". " + y +" (time: " + str(abs(int(self.__duration_db[y]))) + " mins)")
            counter += 1

        print("\nAbsent: " + str(len(self.__report["absent"])))
        counter = 1
        for y in self.__report['absent']:
            print("\t" + str(counter) + ". " + y)
            counter += 1

        print("\nPotential Errors (spelling mistakes, etc.): " + str(len(self.__error_db[0])))
        counter = 1
        for y in range(len(self.__error_db[0])):
            print("\t" + str(counter) + ". " + self.__error_db[0][y] + " (replaced with " + self.__error_db[1][y] + ")")
            counter += 1
        print()