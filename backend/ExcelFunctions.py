import string
from openpyxl import load_workbook  # WORKS ONLY WITH CONDA, use command line - pip install openpyxl

def alphabet_2_number(alpha):
    """
    Converts Excel column identifier to number (eg. AB -> 28) [ Helper for cell_2_row_col() ]
    :analysis: O(log(n))
    :param alpha: an uppercase-only string
    :return: a positive integer
    """
    case = string.ascii_uppercase
    alpha = alpha[::-1]  # Reverses the string
    number = 0
    count = 0
    for x in alpha:
        number += (case.index(x) + 1) * 26 ** count
        count += 1
    return number


def number_2_alphabet(number):
    """
    Converts a row index to Excel column identifier (eg. 28 -> AB) [ Helper for row_col_2_cell() ]
    :analysis: O(log (n))
    :param number: a positive integer
    :return: an uppercase-only string
    """
    case = string.ascii_uppercase
    alpha = ""
    while number >= 26:
        alpha += case[number % 26 - 1]
        number = number // 26
    alpha += case[number - 1]
    return alpha[::-1]


def column_generator(limit):
    """
    Creates a list of Excel column identifiers (in the actual order) (Eg. A, B, ... , Z, AA, AB, ..., ZZZ)
    :analysis: O(n)
    :param limit: positive integer
    :return: a list of uppercase-only strings
    """
    final = list()
    case = list(string.ascii_uppercase)
    case.insert(0, "")
    count = 0
    for x in range(len(case)):
        for y in range(len(case)):
            for z in range(1, len(case)):
                final.append(str(case[x]) + str(case[y]) + str(case[z]))
                count += 1
                if count == limit:
                    break
            if count == limit:
                break
        if count == limit:
            break
    return final


def row_col_2_cell(row, col):
    """
    Converts Cartesian cell representation (row, col) to Excel cell representation (eg. (1, 2) -> B1)
    :analysis: O(1)
    :param col: positive integer
    :param row: positive integer
    :return: a string
    """
    return number_2_alphabet(col) + str(row)


def cell_2_row_col(cell):
    """
    Converts Excel cell to Cartesian cell representation (row, col) (Eg. B1 -> (1, 2))
    :analysis: O(n)
    :param cell: a string
    :return: tuple(positive integer, positive integer)
    """
    count = 0
    while not cell[count].isdigit():
        count += 1
    return int(cell[count:]), alphabet_2_number(cell[:count])


if __name__ == '__main__':
    file = None  # file name (incl. extension)
    if file is not None:
        workbook = load_workbook(file)  # File opening command
        print("Sheets in workbook-", workbook.sheetnames)  # Print names of all the sheets in the workbook

        workbook._active_sheet_index = 0  # Set current sheet index
        sheet = workbook.active  # Gets reference to the currently active sheet in the workbook, unless overridden
        print("Current sheet name =", sheet.title)  # Prints sheet name

        cell_sheet = "B1"  # cell to fetch data from

        cell_sheet = cell_2_row_col(cell_sheet)
        print("Value in cell", str(cell_sheet), "=",
              sheet.cell(cell_sheet[0], cell_sheet[1]).value)  # Gets value from the given cell
