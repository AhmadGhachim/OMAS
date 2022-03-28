""" WARNING: RUNNING THIS MAY RESET ANY EXISTING WORKBOOK(S) WITH THE SAME NAME"""
import os
import openpyxl
from openpyxl.styles import Font

path = os.path.abspath(os.path.pardir)
path = path.replace("\\", "/", path.count("\\"))

def workbook_initializer(class_name: str, start_end_month: tuple, class_list_path: str):
    """
    Creates an empty workbook with student names, with a sheet for each month.
    (min course duration-> 1 month; max course duration-> 12 months)
    :param class_name: string representing name of the class
    :param start_end_month: tuple containing course duration (start_month: str, end_month: str)
    :param class_list_path: path to class list (text file containing names of students)
    """
    original_months = ['January', 'February', 'March', 'April', 'May', 'June',
                       'July', 'August', 'September', 'October', 'November', 'December']

    # If the course ends next year
    if original_months.index(start_end_month[0]) > original_months.index(start_end_month[1]):
        months = original_months[original_months.index(start_end_month[0]):] + original_months[:original_months.index(start_end_month[1]) + 1]

    # If the course starts & ends in the same year
    else:
        months = original_months[original_months.index(start_end_month[0]):original_months.index(start_end_month[1]) + 1]

    class_list_copy = list()

    try:
        with open(class_list_path, encoding="utf-8", errors="strict") as f:
            class_list = f.read()
            class_list = list(filter(None, class_list.split("\n")))
            for name in class_list:
                name = name.strip()
                class_list_copy.append(name)
    except UnicodeError:
        with open(class_list_path, encoding="utf-16", errors="strict") as f:
            class_list = f.read()
            class_list = list(filter(None, class_list.split("\n")))
            for name in class_list:
                name = name.strip()
                class_list_copy.append(name)

    class_list_copy.sort()

    wb = openpyxl.Workbook()
    wb.active.title = months[0]
    for month in months[1:]:
        wb.create_sheet(month)

    for month in wb.sheetnames:
        wb._active_sheet_index = wb.sheetnames.index(month)
        sheet = wb.active
        sheet["A1"] = "NAME"
        ft = Font(bold=True)
        sheet["A1"].font = ft
        count = 2
        for z in class_list_copy:
            sheet["A" + str(count)] = z.upper() # Names are stored in uppercase to ease spelling correction
            count += 1
    wb._active_sheet_index = 0
    try:
        wb.save(path + "/backend/Excel files/" + class_name + ".xlsx")
    except FileNotFoundError:
        os.mkdir(path + "/backend/Excel files")
        wb.save(path + "/backend/Excel files/" + class_name + ".xlsx")

def add_student(class_name: str, student_name: str):
    """
    Adds a new student entry to the attendance database
    :param class_name: string representing the name of the class
    :param student_name: string representing the name of the student
    """
    student_name = student_name.upper()
    names = list()
    wb = openpyxl.load_workbook("Excel files/" + class_name + ".xlsx")  # Open class workbook
    for x in range(len(wb.sheetnames)):
        wb._active_sheet_index = x
        sheet = wb.active
        # Create a list of student names ('names') from the workbook
        row_count = 2
        cell_data = sheet.cell(row_count, 1).value
        while cell_data is not None:
            names.append(str(cell_data).upper())
            row_count += 1
            cell_data = sheet.cell(row_count, 1).value

        index = 0
        while min(names[index], student_name) != student_name:
            index += 1

        sheet.insert_rows(index + 2, 1)
        sheet.cell(index + 2, 1).value = student_name
    wb.save("Excel files/" + class_name + ".xlsx")
    wb.close()

if __name__ == '__main__':
    print("Testing WorkbookInitializer")
    print("\tInitializing data...")
    test_classes = [x[0:-4] for x in os.listdir("Class List") if x[-4:] == ".txt"]
    test_start_end_months = [("January", "April"), ("September", "January"), ("July", "June")]
    test_class_lists = ["Class List/" + x + ".txt" for x in test_classes]

    print("\tRunning function...")
    for class_name_item, start_end, class_list_item in zip(test_classes, test_start_end_months, test_class_lists):
        workbook_initializer(class_name_item, start_end, class_list_item)

    print("\tFiles created successfully. Verifying workbook format...")
    test_duration = [['January', 'February', 'March', 'April'], ['September', 'October', 'November', 'December', 'January'],
                     ['July', 'August', 'September', 'October', 'November', 'December', 'January', 'February', 'March',
                      'April', 'May', 'June']]

    wb1 = openpyxl.load_workbook(path + "/backend/Excel files/CMPT 370.xlsx")
    assert wb1.sheetnames == test_duration[1], "Sheet range incorrect (months in the same year)"

    wb1 = openpyxl.load_workbook(path + "/backend/Excel files/MATH 211.xlsx")
    assert wb1.sheetnames == test_duration[2], "Sheet range incorrect (months in consecutive years)"
    print("\tPassed both tests")

    print("\tDeleting test workbooks...")
    for class_name_item in test_classes:
        os.remove("Excel files/" + class_name_item +".xlsx")
    os.removedirs("Excel files")

    print("** End of testing **")




